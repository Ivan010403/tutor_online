import telebot
import json
import time
import traceback
from datetime import datetime
import threading, os
import openpyxl


path = "/root/go/tg_bot/"
#path = ""
 
mutex = threading.Lock()
mutex_photo = threading.Lock()

bot = telebot.TeleBot('__')

data_users = {} # id: fio, days
data_results = {} #id: [first, second, ..., seventh]
data_days = {} # fio: [6 days]

number_day = 1
special_for_excel = 0

data_time = {}
data_image = {} # name_of_file : file_id

top_by_day = {} #day: [top]

start_message = "Справка по обращению с ботом: \n"\
                "/start - начало работы для новых пользователей\n"\
                "/help - вывести справку по обращению с ботом\n"\
                "/top - вывести топ участников по дням\n"\
                "/edit_name - изменить введённое ФИО\n"\
                "/print_name - вывести введённое ФИО"

def read_from_file():
    f = open(path+"data_results.txt", 'r', encoding='utf-8')
    global data_results
    data_results = json.load(f)
    f.close()

    f = open(path+"data_users.txt", 'r', encoding='utf-8')
    global data_users
    data_users = json.load(f)
    f.close() 

    f = open(path+"top_by_day.txt", 'r', encoding='utf-8')
    global top_by_day
    top_by_day = json.load(f)
    f.close()

    f = open(path+"number_day.txt", 'r', encoding='utf-8')
    global number_day
    number_day = int(f.readline())
    f.close()

    f = open(path+"number_excel.txt", 'r', encoding='utf-8')
    global special_for_excel
    special_for_excel = int(f.readline())
    f.close()

def save_to_file(): 
    f = open(path + "data_users.txt", 'w', encoding='utf-8') # TODO: f = open("/root/go/tg_bot/data_users.txt", 'w')
    global data_users
    json.dump(data_users, f)
    f.close()

def save_to_file_data_results():
    f = open(path + "data_results.txt", 'w', encoding='utf-8') # TODO: f = open("/root/go/tg_bot/data_results.txt", 'w')
    global data_results
    json.dump(data_results, f)
    f.close()

def save_to_file_number_day():
    f = open(path + "number_day.txt", 'w', encoding='utf-8') # TODO: f = open("/root/go/tg_bot/data_results.txt", 'w')
    global number_day
    f.write(str(number_day))
    f.close()

def save_top_to_file():
    f = open(path + "top_by_day.txt", 'w', encoding='utf-8')
    global top_by_day
    json.dump(top_by_day, f)
    f.close()

def send_logs(text):
    f = open(path + "log.txt", 'a', encoding='utf-8') #/root/go/tg_bot/log.txt
    f.write(f'{datetime.now().strftime("%H:%M:%S")} ' + text)
    f.close()

def upload_data_from_tests():
    for i in range (1, number_day):
        f = open(f'{path}tests/day{i}.txt', 'r', encoding='utf-8')
        lines = [line.rstrip('\n') for line in f]
        f.close()

        for l in lines:
            if l != '':
                if l not in data_days:
                    data_days[l] = [0,0,0,0,0,0]
                    # сделать так чтобы не переполнялся за 1 (если уже есть 1 не надо инкрементить дальше)
                if i == 1:
                    data_days[l][0] +=1
                if i == 2:
                    data_days[l][1] +=1
                if i == 3:
                    data_days[l][2] +=1
                if i == 4:
                    data_days[l][3] +=1
                if i == 5:
                    data_days[l][4] +=1
                if i == 6:
                    data_days[l][5] +=1

    send_logs(f'Метод upload_data_from_tests() отработал. Данные по дням: \n data_days: {data_days}\n')

def send_needed_data():
    for l in data_users:
        if data_users[l][0] in data_days:

            if data_users[l][1]-1 <= 5 and data_days[data_users[l][0]][data_users[l][1]-1]>=1:
                print(data_users[l][0])
                data_users[l][1] += 1

                f = open(f'{path}lessons/day{data_users[l][1]}.txt', 'r', encoding='utf-8')
                links = [line.rstrip('\n') for line in f]
                strin = "\n".join(links)
                bot.send_message(int(l), f'Ваши лекции {data_users[l][1]} дня: \n{strin}')
                f.close()

                send_logs(f'{data_users[l][0]} получил лекции {data_users[l][1]} дня\n')

                mutex.acquire()
                save_to_file()
                mutex.release()

def get_data():
    for i in range (1, number_day):
        os.system(f'python3 /root/go/tg_bot/mm.py {i}')
        bot.send_message(-4190522872, f'Запрос к таблице по тестам {i} дня завершён') 
        send_logs(f'Запрос к таблице по тестам {i} дня завершён\n')

    upload_data_from_tests()

    send_needed_data()

    bot.send_message(-4190522872, f'Работа метода get_data() завершена! Подробнее в log.txt\n') 


    t = threading.Timer(3600, get_data) #TODO: изменить частоту !!!
    t.daemon = True
    t.start()

try:
    bot.send_message(-4190522872, "<b>Бот запустился, начинается чтение файлов</b>", parse_mode='html')
    send_logs(f'Бот запустился. Начинается чтение файлов\n')

    read_from_file()

    bot.send_message(-4190522872, f'Чтение файлов успешно\n\n ДЕНЬ = {number_day} \n\n ЭКСЕЛЬ = {special_for_excel}')
    send_logs(f'Чтение файлов успешно!\n\n data_users = {data_users} \n\n data_results = {data_results}')
except Exception as e:
    exception_traceback = traceback.format_exc()
    bot.send_message(-4190522872, f'Ошибка при выполнении метода read_from_file():\n{exception_traceback}') 
    send_logs(f'{exception_traceback}\n')

try:
    get_data()
except Exception as e:
    exception_traceback = traceback.format_exc()
    bot.send_message(-4190522872, f'Ошибка при выполнении метода get_data():\n\n{exception_traceback}') 
    send_logs(f'{exception_traceback}\n')

#--------------работа с топами и сортировками------------
def get_data_from_excel():
    temp = {}
    global val1, val2

    wookbook = openpyxl.load_workbook(f'{path}results_of_tests/day{special_for_excel}.xlsx')
    worksheet = wookbook.active

    for i in range(0, worksheet.max_row):
        for col1 in worksheet.iter_cols(1, 1):
            val1 = col1[i].value
            if val1[len(val1)-1] == " ":
                val1 = val1[:len(val1)-1]
        for col2 in worksheet.iter_cols(2, 2):
            val2 =  col2[i].value

        temp[val1] = val2
    
    for l in data_results:
        if data_users[l][0] in temp:
            data_results[l][special_for_excel-1] = temp[data_users[l][0]]
    
    mutex.acquire()
    save_to_file_data_results()
    mutex.release()

    send_logs(f'Загружена табличка эксель под именем day{special_for_excel}.xlsx и выгружены данные. Сейчас данные равны data_results = {data_results}\n')

def get_top_by_day():
    for i in range (1, special_for_excel+1):
        temp = []

        for l in data_results:
            temp.append([l, data_results[l][i-1]])
        
        temp.sort(key=lambda student: student[1], reverse=True)

        top_by_day[i] = []

        top_by_day[i] = temp[0:8]
#-----------------------служебное------------------------

def time_check(message):
    current_time = datetime.now()
    last_datetime = data_time.get(message.from_user.id)

    # Если первое сообщение (время не задано)
    if not last_datetime:
        data_time[message.from_user.id] = current_time
    else:
        delta_seconds = (current_time - last_datetime).total_seconds()
        
        # Если время ожидания не закончилось
        if delta_seconds < 2:
            bot.send_message(message.from_user.id, f'Не спеши! Подожди пару секунд и напиши снова!')
            data_time[message.from_user.id] = current_time
            return True
        
        data_time[message.from_user.id] = current_time
        return False

def send_photo(id, name):
    if f'{name}.jpg' not in data_image:
        mutex_photo.acquire()

    if f'{name}.jpg' not in data_image:
        f = open(f'{path}image/{name}.jpg', "rb") #/root/go/tg_bot/image/
        msg = bot.send_photo(chat_id=id, photo=f) 
        f.close()

        data_image[f'{name}.jpg'] = msg.photo[0].file_id
        if mutex_photo.locked():
            mutex_photo.release()   
    else:
        if mutex_photo.locked():
            mutex_photo.release()
        bot.send_photo(id, photo=data_image[f'{name}.jpg'])

try:
    @bot.message_handler(commands=['start'])
    def start_handler(message):
        if time_check(message):
            return
        
        print(str(message.from_user.id))
        #print(data_users[str(message.from_user.id)])
        if str(message.from_user.id) not in data_users:

            send_photo(message.from_user.id, "hello")
            bot.send_message(message.from_user.id, "Приветствуем тебя, дорогой участник \"Школы Тьюторов\"! Введи своё ФИО!", parse_mode='html')

            data_users[str(message.from_user.id)] = ["first", 1]

            mutex.acquire()
            save_to_file()
            mutex.release()

    @bot.message_handler(commands=['help'])
    def help_handler(message):
        if time_check(message):
            return
        bot.send_message(message.from_user.id, start_message)
    
    @bot.message_handler(commands=['top'])
    def help_handler(message):
        if time_check(message):
            return
        if str(message.from_user.id) not in data_users:
            bot.send_message(message.from_user.id, "Вы ещё не зарегистрированы!")
            return
        
        text = ""
        for l in top_by_day:
            text += f'Топ участников по баллам за тестирование {l} дня:\n'
            for i in range (0, len(top_by_day[l])):
                if top_by_day[l][i][0] in data_users:
                    text += f'{i+1}) {data_users[top_by_day[l][i][0]][0]} с {top_by_day[l][i][1]} баллами\n'
            text +="\n"
        
        if text == "":
            bot.send_message(message.from_user.id, "Рейтинга по дням ещё нет (тест прошли слишком мало людей)")
        else:
            bot.send_message(message.from_user.id, text)

    
    @bot.message_handler(commands=['edit_name'])
    def help_handler(message):
        if time_check(message):
            return
        if str(message.from_user.id) not in data_users:
            bot.send_message(message.from_user.id, "Вы ещё не зарегистрированы!")
            return
        
        bot.send_message(message.from_user.id, "Введите новое ФИО (проверьте ещё раз на наличие ошибок)!")

        data_users[str(message.from_user.id)][0] = "first"

        mutex.acquire()
        save_to_file()
        mutex.release()

    @bot.message_handler(commands=['print_name'])
    def print_name(message): 
        if time_check(message):
            return
        if str(message.from_user.id) not in data_users:
            bot.send_message(message.from_user.id, "Вы ещё не зарегистрированы!")
            return
        
        bot.send_message(message.from_user.id, f'Ваше ФИО на данный момент = "{data_users[str(message.from_user.id)][0]}"')

    # --------------сверху - ручки!-----------------------
    
    @bot.message_handler(content_types=['text'])
    def get_text_messages(message):
        if time_check(message):
            return
        
        if str(message.from_user.id) in data_users and data_users[str(message.from_user.id)][0] == "first":
            data_users[str(message.from_user.id)][0] = message.text
            
            if str(message.from_user.id) not in data_results:
                data_results[str(message.from_user.id)] = [0, 0, 0, 0, 0, 0]

            mutex.acquire()
            text = f'Пользователь {data_users[str(message.from_user.id)][0]} с id {message.from_user.id} успешно зарегистрировался\n'
            send_logs(text)
            mutex.release()

            bot.send_message(message.from_user.id, f'Вы успешно зарегистрировались! {start_message}')

            f = open(f'{path}lessons/day{data_users[str(message.from_user.id)][1]}.txt', 'r', encoding='utf-8')
            links = [line.rstrip('\n') for line in f]
            strin = "\n".join(links)
            bot.send_message(message.from_user.id, f'Ваши лекции {data_users[str(message.from_user.id)][1]} дня: \n{strin}')
            f.close()
            
            mutex.acquire()
            save_to_file()
            save_to_file_data_results()
            mutex.release()
            
        elif message.text =="/remove_all":
            data_users.clear()
            data_results.clear()
            global number_day
            global special_for_excel
            number_day = 1

            save_to_file()
            save_to_file_data_results()
            save_to_file_number_day()
            bot.send_message(-4190522872, f'Данные на текущий момент \n\n ДЕНЬ = {number_day} \n\n ЭКСЕЛЬ = {special_for_excel}')
        
        elif message.text =="/print_data":
            bot.send_message(-4190522872, f'Данные на текущий момент \n\n ДЕНЬ = {number_day} \n\n ЭКСЕЛЬ = {special_for_excel}')

            f = open(path+"data_users.txt", 'rb')            
            bot.send_document(-4190522872, f)
            f.close()

            f = open(path+"data_results.txt", 'rb')            
            bot.send_document(-4190522872, f)
            f.close()
        
        elif message.text =="/next_day":
            number_day += 1
            save_to_file_number_day()
            bot.send_message(-4190522872, f'Текущий максимальный день: <b>{number_day}</b>',parse_mode='html')
        elif message.text =="/send_logs":
            f = open(path+"log.txt", 'rb')            
            bot.send_document(-4190522872, f)
            f.close()

        elif message.text =="/send_excel":
            special_for_excel += 1
            get_data_from_excel()

            f = open(path+"number_excel.txt", 'w', encoding='utf-8') # TODO: f = open("/root/go/tg_bot/data_results.txt", 'w')
            f.write(str(special_for_excel))
            f.close()

            bot.send_message(-4190522872, f'Загрузка эксель файла прошла успешно! special_for_excel = {special_for_excel}',parse_mode='html')
            f = open(path+"data_results.txt", 'rb')            
            bot.send_document(-4190522872, f)
            f.close()
        
        elif message.text =="/put_top_by_day":
            get_top_by_day()
            save_top_to_file()

            bot.send_message(-4190522872, f'Топ по дням высчитан: \n{top_by_day}' ,parse_mode='html')

        elif message.text =="/admin_change_fio":
            send = bot.send_message(message.chat.id, 'Введите id и новое ФИО (id + ПРОБЕЛ + ФИО!)')
            bot.register_next_step_handler(send, change_name)

        elif message.text =="/admin_send":
            send = bot.send_message(message.chat.id, 'Введите id и сообщение')
            bot.register_next_step_handler(send, send_data)
        else:
            bot.send_message(message.chat.id, "Неверный ввод! Попробуйте ещё!")

    def change_name(message):
        l = message.text.split()

        if l[0] in data_users:
            data_users[l[0]][0] = l[1] + " " + l[2] + " " + l[3]
            mutex.acquire()
            save_to_file()
            mutex.release()
        else:
            bot.send_message(message.chat.id, "Неверный ввод")

    def send_data(message):
        l = message.text.splitlines()
        text = ""
        #l[0] = l[0][0:len(l[0])-1]


        if l[0] in data_users:
            for i in range (1, len(l)):
                text += l[i] + "\n"
            bot.send_message(l[0], text)
            bot.send_message(message.from_user.id, "Сообщение отправлено!")
        else:
            bot.send_message(message.chat.id, "Неверный ввод. Вы ввели id " + l[0])


    bot.infinity_polling(timeout=10, long_polling_timeout = 5)

except Exception as e:
    exception_traceback = traceback.format_exc()
    print(exception_traceback)
    bot.send_message(-4190522872, f'{exception_traceback}') 
